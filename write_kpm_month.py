#!/usr/bin/env python3
"""
Write a month's Planning Center KPM row into 'MCC Connection KPMs.xlsx'.

Charts/drawings in this workbook would be dropped by openpyxl.save(), so this
edits the target worksheet's XML in place (surgical string edit inside the .xlsx
zip), leaving every other part byte-for-byte intact. openpyxl is used READ-ONLY
to locate the row/columns.

Writes raw PCO inputs (active, students, kids, households, life-group members,
life/care groups, serving Sundays, serving anywhere) and replicates the sheet's
own formulas for Adults, Total, and the % columns so a human opening the file
sees the same computed columns as prior months.

Usage:
  python3 write_kpm_month.py --file "<path.xlsx>" --tab "2026 KPMs" --month June \\
      --active 1517 --students 155 --kids 333 --households 470 \\
      --life_group_members 384 --life_care_groups 612 \\
      --serving_sundays 387 --serving_anywhere 444
"""
import argparse, io, os, re, shutil, zipfile
import openpyxl

HEADERS = {  # canonical -> lowercased header label (matched by startswith/contains)
    "active": "active profiles", "adults": "active adults", "students": "active students",
    "kids": "active kids", "households": "unique households",
    "life_group_members": "unique mem of a life group",
    "life_care_groups": "life / class / care", "total": "total groups and students",
    "serving_sundays": "serving sundays", "serving_anywhere": "serving anywhere",
    "pct_circle": "% adults in a circle", "pct_circle_all": "% adults and students",
    "pct_serving": "% serving",
}

def col_letter(idx):
    return openpyxl.utils.get_column_letter(idx)

def locate(ws):
    """Return (month_row_number_fn, cols) by scanning the header row for labels."""
    norm = lambda v: str(v).strip().lower() if v is not None else ""
    hdr_row = None
    for row in ws.iter_rows(min_row=1, max_row=200):
        for c in row:
            if norm(c.value) == "active profiles":
                hdr_row = c.row; break
        if hdr_row:
            break
    if not hdr_row:
        raise SystemExit("could not find 'ACTIVE profiles' header")
    cols = {}
    for c in ws[hdr_row]:
        lbl = norm(c.value)
        for key, want in HEADERS.items():
            if lbl == want or lbl.startswith(want) or want in lbl:
                cols.setdefault(key, col_letter(c.column))
    return hdr_row, cols

def month_row(ws, hdr_row, month):
    want = month.strip().lower()[:3]
    for r in range(hdr_row + 1, hdr_row + 14):
        lbl = str(ws.cell(r, 1).value or "").strip().lower()[:3]
        if lbl == want:
            return r
        if lbl not in ("jan","feb","mar","apr","may","jun","jul","aug","sep","oct","nov","dec"):
            break
    raise SystemExit("month row not found for " + month)

def set_cell(xml, ref, inner):
    """Replace the (empty or filled) cell element for `ref`, preserving its attributes."""
    pat = re.compile(r'(<c r="%s"[^>]*?)(/>|>.*?</c>)' % re.escape(ref), re.S)
    if not pat.search(xml):
        raise SystemExit("cell %s not present in sheet XML" % ref)
    return pat.sub(lambda m: m.group(1) + ">" + inner + "</c>", xml, count=1)

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", required=True); ap.add_argument("--tab", default="2026 KPMs")
    ap.add_argument("--month", required=True)
    for k in ("active","students","kids","households","life_group_members",
              "life_care_groups","serving_sundays","serving_anywhere"):
        ap.add_argument("--" + k, type=int, required=True)
    a = ap.parse_args()

    wb = openpyxl.load_workbook(a.file, read_only=True)
    ws = wb[a.tab]
    hdr_row, cols = locate(ws)
    r = month_row(ws, hdr_row, a.month)
    wb.close()

    B, C, D, E, F = cols["active"], cols["adults"], cols["students"], cols["kids"], cols["households"]
    G, H, I = cols["life_group_members"], cols["life_care_groups"], cols["total"]
    J, K = cols["serving_sundays"], cols["serving_anywhere"]
    M, N, O = cols["pct_circle"], cols["pct_circle_all"], cols["pct_serving"]

    # raw values
    writes = {
        f"{B}{r}": ("n", a.active), f"{D}{r}": ("n", a.students), f"{E}{r}": ("n", a.kids),
        f"{F}{r}": ("n", a.households), f"{G}{r}": ("n", a.life_group_members),
        f"{H}{r}": ("n", a.life_care_groups), f"{J}{r}": ("n", a.serving_sundays),
        f"{K}{r}": ("n", a.serving_anywhere),
        # formulas (match existing rows' patterns; offsets: student block = row+35, attendance C = row-36)
        f"{C}{r}": ("f", f"({B}{r}-({D}{r}+{E}{r}))"),
        f"{I}{r}": ("f", f"{H}{r}+E{r+35}"),
        f"{M}{r}": ("f", f"({H}{r}/{C}{r})"),
        f"{N}{r}": ("f", f"{H}{r}/C{r-36}"),
        f"{O}{r}": ("f", f"{K}{r}/C{r-36}"),
    }

    # locate sheet xml part
    zin = zipfile.ZipFile(a.file)
    wbx = zin.read("xl/workbook.xml").decode("utf8")
    rels = zin.read("xl/_rels/workbook.xml.rels").decode("utf8")
    sid = dict(re.findall(r'<sheet[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"', wbx))[a.tab]
    tgt = re.findall(r'Id="([^"]+)"[^>]*Target="([^"]+)"', rels)
    relmap = {i: t for i, t in tgt}
    sheet_path = "xl/" + relmap[sid].lstrip("/")

    xml = zin.read(sheet_path).decode("utf8")
    for ref, (typ, val) in writes.items():
        inner = f"<v>{val}</v>" if typ == "n" else f"<f>{val}</f>"
        xml = set_cell(xml, ref, inner)

    # rewrite zip: copy every part, swap the one sheet
    tmp = a.file + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename == sheet_path:
                data = xml.encode("utf8")
            zout.writestr(item, data)
    zin.close()
    shutil.move(tmp, a.file)
    print(f"wrote {a.month} (row {r}) to {os.path.basename(a.file)} :: "
          f"{B}{r}={a.active} {D}{r}={a.students} {E}{r}={a.kids} {F}{r}={a.households} "
          f"{G}{r}={a.life_group_members} {H}{r}={a.life_care_groups} {J}{r}={a.serving_sundays} {K}{r}={a.serving_anywhere}; "
          f"formulas set for {C}{r},{I}{r},{M}{r},{N}{r},{O}{r}")

if __name__ == "__main__":
    main()
