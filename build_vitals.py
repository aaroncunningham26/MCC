#!/usr/bin/env python3
"""
MCC Vitals Dashboard renderer.

Reads the source-of-truth spreadsheet ...
  Data Sheets/Data Source Sheets/MCC VITALS.xlsx  (tab: VITALS)
... and writes:
  MCC/data/vitals.json   -- structured data (years, metrics, values, pcts, through-week)
  MCC/vitals.html        -- the leadership dashboard, numbers baked in

The VITALS tab lays each year out as two columns: a value column and a
second column (percent for most metrics, per-person $ for Giving, YoY growth %
for the attendance rows). Column pairs:
   2022 D/E   2023 F/G   2024 H/I   2025 J/K   2026 L/M
Row 26 col C holds the current "Thru Week #".

Monthly refresh: update the spreadsheet (or run build_vitals_pull.py to pull the
current numbers from Planning Center), then run:  python3 build_vitals.py
"""
import json, os, datetime
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.normpath(os.path.join(
    BASE, "..", "Data Sheets", "Data Source Sheets", "MCC VITALS.xlsx"))
YEARS = [2022, 2023, 2024, 2025, 2026]
VAL_COLS = {2022: 4, 2023: 6, 2024: 8, 2025: 10, 2026: 12}   # D F H J L
PCT_COLS = {2022: 5, 2023: 7, 2024: 9, 2025: 11, 2026: 13}   # E G I K M

# row -> metric definition. `pct_is` tells us what the 2nd column means.
METRICS = [
    # Evangelism
    dict(row=2,  group="evangelism",  name="Avg. attendance (in-person)", goal="Min. 5% growth",
         type="growth", goalMin=5, isAvg=True,  pct_is="growth"),
    dict(row=3,  group="evangelism",  name="Avg. adults (in-person)",     goal="Min. 5% growth",
         type="growth", goalMin=5, isAvg=True,  pct_is="growth"),
    dict(row=4,  group="evangelism",  name="Avg. students (in-person)",   goal="10%–15% of attendance",
         type="range",  goalMin=10, goalMax=15, isAvg=True, pct_is="share"),
    dict(row=5,  group="evangelism",  name="Avg. kids (in-person)",       goal="15%–25% of attendance",
         type="range",  goalMin=15, goalMax=25, isAvg=True, pct_is="share"),
    dict(row=6,  group="evangelism",  name="Avg. online attendance",      goal="25%–30% of attendance",
         type="range",  goalMin=25, goalMax=30, isAvg=True, pct_is="share"),
    dict(row=7,  group="evangelism",  name="Baptisms / POF",              goal="10% of attendance",
         type="min",    goalMin=10, isAvg=False, pct_is="share"),
    dict(row=8,  group="evangelism",  name="Sunday visitors",             goal="100% of attendance",
         type="min",    goalMin=100, isAvg=False, pct_is="share"),
    dict(row=9,  group="evangelism",  name="Connect breakfast",           goal="10% of attendance",
         type="min",    goalMin=10, isAvg=False, pct_is="share"),
    # Discipleship
    dict(row=16, group="discipleship", name="Adults in circles",      goal="70% of adults",
         type="min",   goalMin=70, isAvg=True,  pct_is="share"),
    dict(row=17, group="discipleship", name="Regular serving",        goal="50% of attendance",
         type="min",   goalMin=50, isAvg=True,  pct_is="share"),
    dict(row=18, group="discipleship", name="Unique donors",          goal="40%–60% of attendance",
         type="range", goalMin=40, goalMax=60, isAvg=True, pct_is="share"),
    dict(row=19, group="discipleship", name="New donors",             goal="5%–10% of unique donors",
         type="range", goalMin=5, goalMax=10, isAvg=False, pct_is="share"),
    dict(row=20, group="discipleship", name="Giving / person / week",  goal="$25–$35 per person/week",
         type="dollar", goalMin=25, goalMax=35, isAvg=False, pct_is="dollar"),
    dict(row=21, group="discipleship", name="Starting Point",         goal="10% of guests",
         type="min",   goalMin=10, isAvg=False, pct_is="share"),
]

def num(v):
    if v is None: return None
    if isinstance(v, str):
        try: return float(v)
        except ValueError: return None
    return float(v)

def extract():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["VITALS"]
    # through-week
    through_week = None
    for r in range(1, ws.max_row + 1):
        if str(ws.cell(r, 2).value or "").strip().startswith("Thru Week"):
            through_week = int(num(ws.cell(r, 3).value) or 0); break
    # close-of month from KPM Format tab
    month = None
    kf = wb["KPM Format"]
    for row in kf.iter_rows():
        for c in row:
            if c.value and "close of" in str(c.value).lower():
                month = str(kf.cell(c.row, c.column + 1).value or "").strip()
    out = []
    for m in METRICS:
        r = m["row"]
        raw_val = {y: num(ws.cell(r, VAL_COLS[y]).value) for y in YEARS}
        raw_pct = {y: num(ws.cell(r, PCT_COLS[y]).value) for y in YEARS}
        if m["pct_is"] == "dollar":
            # display value = per-person $ (2nd col); keep the total separately
            values = [round(raw_pct[y]) if raw_pct[y] is not None else None for y in YEARS]
            pcts = None
            totals = [round(raw_val[y]) if raw_val[y] is not None else None for y in YEARS]
        else:
            values = [round(raw_val[y]) if raw_val[y] is not None else None for y in YEARS]
            pcts = [round(raw_pct[y] * 100) if raw_pct[y] is not None else None for y in YEARS]
            totals = None
        rec = dict(name=m["name"], goal=m["goal"], group=m["group"], type=m["type"],
                   goalMin=m["goalMin"], isAvg=m["isAvg"], values=values, pcts=pcts)
        if "goalMax" in m: rec["goalMax"] = m["goalMax"]
        if totals is not None: rec["totals"] = totals
        out.append(rec)
    wb.close()
    return out, through_week, month

def get(metrics, name):
    for m in metrics:
        if m["name"] == name: return m
    return None

def cur(m):  # current-year (2026) value
    return m["values"][-1]
def curpct(m):
    return m["pcts"][-1] if m.get("pcts") else None

def build_insights(metrics):
    att   = get(metrics, "Avg. attendance (in-person)")
    circ  = get(metrics, "Adults in circles")
    serv  = get(metrics, "Regular serving")
    stu   = get(metrics, "Avg. students (in-person)")
    don   = get(metrics, "Unique donors")
    bullets = []
    # 1. Attendance
    prev_max = max(v for v in att["values"][:-1] if v is not None)
    hi = cur(att) >= prev_max
    g = curpct(att)
    lead = "the highest on record" if hi else "steady"
    tail = (f"comfortably above the 5% growth goal." if g >= 5
            else f"positive but under the 5% growth goal.")
    bullets.append(("green" if g >= 5 else "amber",
        f"Weekend attendance is {lead} — <strong>{cur(att):,} average</strong> through the "
        f"reporting window, up {g}% over 2025. Growth is {tail}"))
    # 2. Circles
    bullets.append(("green" if curpct(circ) >= circ["goalMin"] else "green",
        f"Adults in circles continues to climb — <strong>{cur(circ):,} adults ({curpct(circ)}%)</strong> "
        f"are in a circle, the strongest on record. The 70% goal is within reach."))
    # 3. Watch areas (below-goal averages)
    watch = []
    if curpct(att) < 5:  watch.append(f"attendance growth ({curpct(att)}%)")
    if curpct(serv) < serv["goalMin"]: watch.append(f"regular serving ({curpct(serv)}%)")
    if curpct(don) < don["goalMin"]:   watch.append(f"unique donors ({curpct(don)}%)")
    if watch:
        joined = ", ".join(watch[:-1]) + (f", and {watch[-1]}" if len(watch) > 1 else watch[0])
        if len(watch) == 1: joined = watch[0]
        bullets.append(("amber",
            f"The key watch areas are {joined} — each sits below its 2026 goal and is worth "
            f"deliberate leadership attention this half of the year."))
    # 4. Students (worst trend)
    s24, s26 = stu["values"][2], stu["values"][-1]
    p24, p26 = stu["pcts"][2], stu["pcts"][-1]
    bullets.append(("red",
        f"Student ministry is the most concerning trend — average students has fallen from "
        f"<strong>{s24} in 2024 to {s26} in 2026</strong> ({p24}% → {p26}% of attendance), "
        f"consistently under the 10–15% goal and still declining."))
    # 5. Donors
    d22 = don["pcts"][0]
    bullets.append(("amber",
        f"Unique donors ({cur(don):,}, {curpct(don)}%) remain well below the 40–60% goal and have "
        f"contracted from {d22}% of attendance in 2022. Track this alongside the finance dashboard "
        f"as a long-term giving-health signal."))
    # 6. Methodology note
    bullets.append(("blue",
        "Count-based metrics (visitors, baptisms, Connect Breakfast, giving) are year-to-date and "
        "should not be compared directly to prior full-year totals. Averages (attendance, serving, "
        "circles) are fully comparable."))
    return bullets

# ── HTML template ──────────────────────────────────────────────────────────
TEMPLATE = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta name="robots" content="noindex, nofollow">
<title>MCC Vitals Dashboard</title>
<link rel="icon" type="image/png" href="favicon.png">
<link rel="apple-touch-icon" href="favicon.png">
<link rel="preconnect" href="https://fonts.googleapis.com">
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Lato:wght@400;700;900&display=swap" rel="stylesheet">
<style>
  :root {
    --slate: #343A44; --bg: #F2F2F2; --card: #ffffff; --ink: #23272e; --muted: #6b7280;
    --line: #e3e5ea; --red: #DC2626; --green: #2F8F5B; --amber: #C8842A; --chart: #4C5564;
  }
  *, *::before, *::after { box-sizing: border-box; margin: 0; padding: 0; }
  body { font-family: 'Lato', system-ui, -apple-system, Arial, sans-serif; background: var(--bg); color: var(--ink); font-size: 15px; line-height: 1.6; }
  .mcc-header { position: relative; overflow: hidden; background: var(--slate); color: #fff; }
  .mcc-header-inner { position: relative; max-width: 1180px; margin: 0 auto; padding: 26px 30px; display: flex; justify-content: space-between; align-items: center; flex-wrap: wrap; gap: 14px; }
  .mcc-header::before { content: ""; position: absolute; inset: 0; background: url('https://storage1.snappages.site/VCFHFT/assets/images/21064405_1024x1024_2500.png'); background-size: 620px; opacity: .065; pointer-events: none; }
  .mcc-header .brand { display: flex; align-items: center; gap: 20px; position: relative; }
  .mcc-header .logo { height: 42px; width: auto; display: block; }
  .mcc-header .back { color: #aeb6c2; font-size: 11px; font-weight: 700; letter-spacing: .12em; text-transform: uppercase; text-decoration: none; display: inline-block; margin-bottom: 5px; }
  .mcc-header .back:hover { color: #fff; }
  .mcc-header h1 { font-size: 27px; font-weight: 900; text-transform: uppercase; letter-spacing: -.02em; line-height: 1; }
  .mcc-header .sub { font-size: 11.5px; font-weight: 400; letter-spacing: .05em; text-transform: uppercase; opacity: .72; margin-top: 6px; }
  .mcc-header .right { position: relative; font-size: 11px; font-weight: 700; letter-spacing: .08em; text-transform: uppercase; opacity: .8; text-align: right; line-height: 1.7; }
  .page { max-width: 1180px; margin: 0 auto; padding: 28px 30px 60px; }
  .section-label { font-size: 11px; font-weight: 700; letter-spacing: .12em; text-transform: uppercase; color: var(--muted); margin-bottom: 12px; margin-top: 32px; }
  .cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 12px; margin-bottom: 4px; }
  .card { background: var(--card); border-radius: 10px; padding: 18px 20px; border: 1px solid var(--line); box-shadow: 0 1px 4px rgba(20,30,60,.05); }
  .card .label { font-size: 11px; color: var(--muted); font-weight: 700; text-transform: uppercase; letter-spacing: .08em; margin-bottom: 6px; }
  .card .value { font-size: 26px; font-weight: 900; color: var(--slate); letter-spacing: -0.5px; }
  .card .sub   { font-size: 12px; margin-top: 4px; }
  .card .sub.green  { color: var(--green); } .card .sub.yellow { color: var(--amber); } .card .sub.red { color: var(--red); }
  .card.top-green  { border-top: 4px solid var(--green); } .card.top-yellow { border-top: 4px solid var(--amber); }
  .card.top-red    { border-top: 4px solid var(--red); }  .card.top-blue   { border-top: 4px solid var(--slate); }
  .insights { background: #F4F5F7; border: 1px solid var(--line); border-radius: 10px; padding: 20px 24px; margin-top: 6px; }
  .insights h2 { font-size: 13px; font-weight: 900; text-transform: uppercase; letter-spacing: .08em; color: var(--slate); margin-bottom: 12px; }
  .insights ul { list-style: none; }
  .insights ul li { padding: 8px 0; border-bottom: 1px solid var(--line); font-size: 13px; color: var(--ink); display: flex; align-items: flex-start; gap: 10px; line-height: 1.55; }
  .insights ul li:last-child { border-bottom: none; }
  .dot { width: 8px; height: 8px; border-radius: 50%; margin-top: 5px; flex-shrink: 0; }
  .dot-green { background: var(--green); } .dot-red { background: var(--red); } .dot-amber { background: var(--amber); } .dot-blue { background: var(--slate); }
  .funnel-wrap { display: flex; align-items: stretch; gap: 0; }
  .funnel-step { flex: 1; display: flex; flex-direction: column; align-items: center; text-align: center; position: relative; }
  .funnel-step:not(:last-child)::after { content: '→'; position: absolute; right: -12px; top: 50%; transform: translateY(-50%); font-size: 20px; color: var(--line); z-index: 1; }
  .funnel-box { width: 100%; background: #F4F5F7; border: 1px solid var(--line); border-radius: 10px; padding: 16px 10px; }
  .funnel-step:nth-child(2) .funnel-box { background: #F0FBF5; border-color: #B6DCCA; }
  .funnel-step:nth-child(3) .funnel-box { background: #FDF5EC; border-color: #E8C99A; }
  .funnel-step:nth-child(4) .funnel-box { background: #F4F5F7; border-color: var(--line); }
  .funnel-num { font-size: 26px; font-weight: 900; color: var(--slate); letter-spacing: -0.5px; }
  .funnel-label { font-size: 12px; font-weight: 700; text-transform: uppercase; letter-spacing: .06em; color: var(--ink); margin-top: 4px; }
  .funnel-conv { font-size: 11px; color: var(--muted); margin-top: 6px; line-height: 1.4; }
  .funnel-conv strong { color: var(--ink); }
  @media (max-width: 600px) { .funnel-wrap { flex-direction: column; gap: 8px; }
    .funnel-step:not(:last-child)::after { content: '↓'; right: auto; top: auto; bottom: -16px; left: 50%; transform: translateX(-50%); } }
  .chart-card { background: var(--card); border-radius: 10px; padding: 22px 22px 18px; border: 1px solid var(--line); margin-bottom: 14px; box-shadow: 0 1px 4px rgba(20,30,60,.05); }
  .chart-card h2 { font-size: 15px; font-weight: 900; text-transform: uppercase; letter-spacing: -.01em; color: var(--slate); margin-bottom: 4px; }
  .chart-card .chart-sub { font-size: 13px; color: var(--muted); margin-bottom: 16px; }
  .chart-wrap { position: relative; width: 100%; }
  .legend { display: flex; flex-wrap: wrap; gap: 14px; margin-bottom: 14px; }
  .legend-item { display: flex; align-items: center; gap: 6px; font-size: 12px; color: var(--muted); font-weight: 700; }
  .legend-dot { width: 10px; height: 10px; border-radius: 2px; flex-shrink: 0; }
  .metrics-table { width: 100%; border-collapse: collapse; font-size: 13px; }
  .metrics-table th { background: #F9FAFB; font-weight: 700; text-transform: uppercase; letter-spacing: .06em; font-size: 11px; color: var(--muted); padding: 9px 12px; text-align: right; border-bottom: 2px solid var(--line); white-space: nowrap; }
  .metrics-table th:first-child { text-align: left; width: 200px; }
  .metrics-table th.goal-col { text-align: left; font-weight: 500; }
  .metrics-table td { padding: 9px 12px; border-bottom: 1px solid #F3F4F6; text-align: right; color: var(--ink); vertical-align: middle; }
  .metrics-table td:first-child { text-align: left; font-weight: 700; color: var(--slate); }
  .metrics-table td.goal-cell { text-align: left; font-size: 11px; color: #9CA3AF; white-space: nowrap; }
  .metrics-table tr:hover td { background: #FAFAFA; }
  .cur-val { font-weight: 900; color: var(--slate); }
  .status { display: inline-flex; align-items: center; gap: 4px; font-size: 11px; font-weight: 700; text-transform: uppercase; letter-spacing: .06em; padding: 3px 8px; border-radius: 12px; white-space: nowrap; }
  .status-green { background: #DCFCE7; color: #15803D; } .status-yellow { background: #FEF9C3; color: #A16207; }
  .status-red { background: #FEE2E2; color: var(--red); } .status-gray { background: #F3F4F6; color: var(--muted); }
  .section-divider { margin: 32px 0 20px; padding-bottom: 10px; border-bottom: 2px solid var(--slate); display: flex; align-items: center; gap: 10px; }
  .section-divider h2 { font-size: 15px; font-weight: 900; text-transform: uppercase; letter-spacing: -.01em; color: var(--slate); }
  .section-divider .badge { font-size: 11px; font-weight: 700; color: var(--muted); background: #F0F2F5; padding: 2px 8px; border-radius: 4px; text-transform: uppercase; letter-spacing: .06em; }
  .ytd-note { background: #FEF9C3; border: 1px solid #FDE68A; border-radius: 8px; padding: 10px 14px; font-size: 12px; color: #92400E; margin-bottom: 14px; }
  .mcc-footer { color: var(--muted); font-size: 11.5px; margin-top: 44px; line-height: 1.8; text-align: center; }
</style>
</head>
<body>
<script>
// ══════════════════════════════════════════════════════════════
//  DATA — generated by build_vitals.py from MCC VITALS.xlsx. Do not hand-edit.
// ══════════════════════════════════════════════════════════════
const VITALS = __DATA__;
const LAST_UPDATED = "__LAST_UPDATED__";
const YEARS = VITALS.years;
const evangelism   = VITALS.metrics.filter(m => m.group === "evangelism");
const discipleship = VITALS.metrics.filter(m => m.group === "discipleship");
</script>

<header class="mcc-header">
  <div class="mcc-header-inner">
  <div class="brand">
    <img class="logo" src="https://storage1.snappages.site/VCFHFT/assets/images/22645050_962x358_500.png" alt="Maple City Chapel">
    <div>
      <a class="back" href="index.html">← Leadership Portal</a>
      <h1>MCC Vitals</h1>
      <div class="sub">Attendance, Engagement &amp; Discipleship Indicators</div>
    </div>
  </div>
  <div class="right">Data through<br>__DATA_THROUGH__</div>
  </div>
</header>

<div class="page">
  <div class="section-label">2026 at a glance</div>
  <div class="cards" id="snapshot-cards"></div>

  <div class="section-label">Key takeaways</div>
  <div class="insights">
    <h2>What this data is telling us</h2>
    <ul id="insights-list">__INSIGHTS__</ul>
  </div>

  <div class="section-label">Attendance trends</div>
  <div class="chart-card">
    <h2>Average weekly attendance — 2022 to 2026</h2>
    <p class="chart-sub">In-person total and online attendance each year. 2026 figures are averages through __THROUGH_SHORT__.</p>
    <div class="legend">
      <span class="legend-item"><span class="legend-dot" style="background:#343A44;"></span>In-person (total)</span>
      <span class="legend-item"><span class="legend-dot" style="background:#8899AA;"></span>Online</span>
    </div>
    <div class="chart-wrap" style="height:220px;"><canvas id="attendanceChart" role="img" aria-label="Weekly attendance trend 2022 to 2026"></canvas></div>
  </div>

  <div class="chart-card">
    <h2>In-person breakdown by group</h2>
    <p class="chart-sub">Average weekly adults, students, and kids — each group's contribution to in-person attendance over time.</p>
    <div class="legend">
      <span class="legend-item"><span class="legend-dot" style="background:#343A44;"></span>Adults</span>
      <span class="legend-item"><span class="legend-dot" style="background:#2F8F5B;"></span>Kids</span>
      <span class="legend-item"><span class="legend-dot" style="background:#DC2626;"></span>Students</span>
    </div>
    <div class="chart-wrap" style="height:220px;"><canvas id="breakdownChart" role="img" aria-label="In-person breakdown adults students kids 2022 to 2026"></canvas></div>
  </div>

  <div class="section-divider"><h2>Evangelism vitals</h2><span class="badge">Outreach &amp; growth</span></div>
  <div class="ytd-note">⚠️ <strong>Note:</strong> Metrics marked † are cumulative counts (not averages). Their 2026 figures reflect year-to-date through __THROUGH_SHORT__ and should not be compared directly to prior full-year totals.</div>
  <div class="chart-card" style="padding: 0; overflow: hidden;"><table class="metrics-table" id="evangelismTable"></table></div>

  <div class="section-divider" style="margin-top: 36px;"><h2>Discipleship vitals</h2><span class="badge">Engagement &amp; generosity</span></div>
  <div class="ytd-note">⚠️ <strong>Note:</strong> New donors, giving, and Starting Point figures (†) are YTD through __THROUGH_SHORT__. Per-person weekly giving reflects the YTD average and rises as the year progresses.</div>
  <div class="chart-card" style="padding: 0; overflow: hidden;"><table class="metrics-table" id="discipleshipTable"></table></div>

  <div class="section-label" style="margin-top: 28px;">Discipleship trends</div>
  <div class="chart-card">
    <h2>Adults in circles and regular serving — % of adults/attendance</h2>
    <p class="chart-sub">Tracking progress toward the 70% circles goal and 50% serving goal.</p>
    <div class="legend">
      <span class="legend-item"><span class="legend-dot" style="background:#343A44;"></span>Adults in circles %</span>
      <span class="legend-item"><span class="legend-dot" style="background:#2F8F5B;"></span>Regular serving %</span>
      <span class="legend-item" style="margin-left:8px; border-left: 1px solid var(--line); padding-left: 12px;"><span style="width:20px; height:2px; background:#343A44; display:inline-block; margin-right:4px; border-top: 2px dashed #343A44; vertical-align: middle;"></span>Circles goal (70%)</span>
      <span class="legend-item"><span style="width:20px; height:2px; background:#2F8F5B; display:inline-block; margin-right:4px; border-top: 2px dashed #2F8F5B; vertical-align: middle;"></span>Serving goal (50%)</span>
    </div>
    <div class="chart-wrap" style="height:220px;"><canvas id="discipleshipChart" role="img" aria-label="Adults in circles and regular serving percentage trend"></canvas></div>
  </div>

  <div class="section-label" style="margin-top: 28px;">Next steps funnel — 2026 YTD</div>
  <div class="chart-card">
    <h2>How people move from visitor to engaged disciple</h2>
    <p class="chart-sub">Each stage shows the YTD count for 2026 and the conversion rate from the previous step. Count metrics are through __THROUGH_SHORT__.</p>
    <div class="funnel-wrap" id="funnel-wrap"></div>
    <p style="font-size:11px; color:var(--muted); margin-top:16px;">† YTD counts through __THROUGH_SHORT__. Adults in circles is a point-in-time average, not a cumulative count — conversion shown vs. Starting Point completers is directional only.</p>
  </div>

  <div class="mcc-footer">
    <p>Maple City Chapel &nbsp;·&nbsp; MCC Vitals Dashboard &nbsp;·&nbsp; Data through __DATA_THROUGH_PLAIN__</p>
    <p style="margin-top:2px;">Source: Planning Center / internal tracking. Updated by the Operations team.</p>
  </div>
</div>

<script src="https://cdnjs.cloudflare.com/ajax/libs/Chart.js/4.4.1/chart.umd.js"></script>
<script>
const CUR = YEARS.length - 1;   // index of current year (2026)
function statusFor(m) {
  const pct = m.pcts ? m.pcts[CUR] : null, val = m.values[CUR];
  if (val === null || val === undefined) return 'gray';
  if (m.type === 'growth') { if (pct >= m.goalMin) return 'green'; if (pct >= m.goalMin - 2) return 'yellow'; return 'red'; }
  if (m.type === 'range')  { if (pct >= m.goalMin && pct <= m.goalMax) return 'green'; if (pct >= m.goalMin - 4) return 'yellow'; return 'red'; }
  if (m.type === 'min')    { if (pct >= m.goalMin) return 'green'; if (pct >= m.goalMin * 0.85) return 'yellow'; return 'red'; }
  if (m.type === 'dollar') { if (val >= m.goalMin && val <= m.goalMax) return 'green'; if (val >= m.goalMin * 0.7) return 'yellow'; return 'red'; }
  return 'gray';
}
const statusLabel = s => ({green:'On target', yellow:'Near target', red:'Below target', gray:'—'}[s]);
function fmtVal(m, i) { const v = m.values[i]; if (v === null || v === undefined) return '—'; if (m.type === 'dollar') return '$' + v; return v.toLocaleString('en-US'); }
function fmtPct(m, i) { if (!m.pcts) return ''; const p = m.pcts[i]; if (p === null || p === undefined) return ''; const s = (m.type === 'growth' ? '+' : '') + p + '%'; return '<span style="font-size:11px;color:#6b7280;">' + s + '</span>'; }

// Snapshot cards
const find = n => VITALS.metrics.find(m => m.name === n);
const snap = [
  { m: find('Avg. attendance (in-person)'), label: 'Avg. weekly attendance', extra: 'Goal: 5% growth' },
  { m: find('Adults in circles'),           label: 'Adults in circles',      extra: 'Goal: 70%' },
  { m: find('Regular serving'),             label: 'Regular serving',        extra: 'Goal: 50%' },
  { m: find('Avg. students (in-person)'),   label: 'Avg. students',          extra: 'Goal: 10–15% of attendance' },
];
document.getElementById('snapshot-cards').innerHTML = snap.map(({m,label,extra}) => {
  const s = statusFor(m), pct = m.pcts ? m.pcts[CUR] : null;
  const sub = pct !== null ? (m.type === 'growth' ? '+'+pct+'% YoY · '+extra : pct+'% · '+extra) : extra;
  return `<div class="card top-${s}"><div class="label">${label}</div><div class="value">${fmtVal(m,CUR)}</div><div class="sub ${s}">${sub}</div></div>`;
}).join('');

// Metrics tables
function buildTable(metrics, id) {
  let html = `<thead><tr><th>Metric</th><th class="goal-col">Goal</th>` +
    YEARS.slice(0,-1).map(y => `<th>${y}</th>`).join('') +
    `<th style="color:#343A44;">${YEARS[CUR]}</th><th>Status</th></tr></thead><tbody>`;
  metrics.forEach(m => {
    const s = statusFor(m), dag = !m.isAvg ? '<span style="color:#9CA3AF;"> †</span>' : '';
    html += `<tr><td>${m.name}${dag}</td><td class="goal-cell">${m.goal}</td>` +
      YEARS.slice(0,-1).map((y,i) => `<td>${fmtVal(m,i)} ${fmtPct(m,i)}</td>`).join('') +
      `<td><span class="cur-val">${fmtVal(m,CUR)}</span> ${fmtPct(m,CUR)}</td>` +
      `<td><span class="status status-${s}">${statusLabel(s)}</span></td></tr>`;
  });
  document.getElementById(id).innerHTML = html + '</tbody>';
}
buildTable(evangelism, 'evangelismTable');
buildTable(discipleship, 'discipleshipTable');

// Funnel
(function() {
  const v = n => find(n).values[CUR];
  const visitors = v('Sunday visitors'), cb = v('Connect breakfast'), sp = v('Starting Point'), ic = v('Adults in circles');
  const c1 = Math.round(cb / visitors * 100), c2 = Math.round(sp / cb * 100);
  const steps = [
    { num: visitors, label: 'Sunday Visitors', conv: 'Starting point — guests who attended a service', color: '#343A44' },
    { num: cb, label: 'Connect Breakfast', conv: `<strong>${c1}%</strong> of visitors attended Connect Breakfast`, color: '#2F8F5B' },
    { num: sp, label: 'Starting Point', conv: `<strong>${c2}%</strong> of Connect Breakfast attendees completed Starting Point`, color: '#C8842A' },
    { num: ic, label: 'Adults in Circles', conv: `<strong>${ic.toLocaleString()}</strong> adults currently in a circle (avg) — the ultimate discipleship goal`, color: '#4C5564' }
  ];
  document.getElementById('funnel-wrap').innerHTML = steps.map(s =>
    `<div class="funnel-step"><div class="funnel-box"><div class="funnel-num" style="color:${s.color}">${s.num.toLocaleString()}</div><div class="funnel-label">${s.label}</div><div class="funnel-conv">${s.conv}</div></div></div>`).join('');
})();

// Charts
Chart.defaults.font.family = "'Lato', system-ui, Arial, sans-serif";
Chart.defaults.font.size = 12; Chart.defaults.color = '#6b7280';
const gc = 'rgba(0,0,0,0.06)', labels = YEARS.map(String);
const vals = n => find(n).values, pctsOf = n => find(n).pcts;

new Chart(document.getElementById('attendanceChart'), {
  type: 'bar',
  data: { labels, datasets: [
    { label: 'In-person', data: vals('Avg. attendance (in-person)'), backgroundColor: '#343A44', borderRadius: 4 },
    { label: 'Online',    data: vals('Avg. online attendance'),      backgroundColor: '#8899AA', borderRadius: 4 } ] },
  options: { responsive: true, maintainAspectRatio: false,
    plugins: { legend: { display: false }, tooltip: { callbacks: { label: ctx => ' ' + ctx.dataset.label + ': ' + ctx.raw } } },
    scales: { x: { grid: { display: false }, ticks: { color: '#9CA3AF' } }, y: { grid: { color: gc }, ticks: { color: '#9CA3AF' }, min: 0 } } }
});
new Chart(document.getElementById('breakdownChart'), {
  type: 'bar',
  data: { labels, datasets: [
    { label: 'Adults',   data: vals('Avg. adults (in-person)'),   backgroundColor: '#343A44' },
    { label: 'Kids',     data: vals('Avg. kids (in-person)'),     backgroundColor: '#2F8F5B' },
    { label: 'Students', data: vals('Avg. students (in-person)'), backgroundColor: '#DC2626', borderRadius: 4 } ] },
  options: { responsive: true, maintainAspectRatio: false,
    plugins: { legend: { display: false }, tooltip: { mode: 'index' } },
    scales: { x: { stacked: true, grid: { display: false }, ticks: { color: '#9CA3AF' } }, y: { stacked: true, grid: { color: gc }, ticks: { color: '#9CA3AF' }, min: 0 } } }
});
new Chart(document.getElementById('discipleshipChart'), {
  type: 'line',
  data: { labels, datasets: [
    { label: 'Adults in circles %', data: pctsOf('Adults in circles'), borderColor: '#343A44', backgroundColor: 'rgba(52,58,68,0.08)', borderWidth: 2.5, pointRadius: 4, tension: 0.3, fill: true },
    { label: 'Regular serving %',   data: pctsOf('Regular serving'),   borderColor: '#2F8F5B', backgroundColor: 'rgba(47,143,91,0.07)', borderWidth: 2.5, pointRadius: 4, tension: 0.3, fill: true },
    { label: 'Circles goal', data: YEARS.map(() => 70), borderColor: '#343A44', borderDash: [5,4], borderWidth: 1.5, pointRadius: 0, fill: false },
    { label: 'Serving goal', data: YEARS.map(() => 50), borderColor: '#2F8F5B', borderDash: [5,4], borderWidth: 1.5, pointRadius: 0, fill: false } ] },
  options: { responsive: true, maintainAspectRatio: false,
    plugins: { legend: { display: false }, tooltip: { callbacks: { label: ctx => ' ' + ctx.dataset.label + ': ' + ctx.raw + '%' } } },
    scales: { x: { grid: { display: false }, ticks: { color: '#9CA3AF' } }, y: { grid: { color: gc }, ticks: { callback: v => v + '%', color: '#9CA3AF' }, min: 0, max: 100 } } }
});
</script>
</body>
</html>
"""

def render(metrics, through_week, month):
    data = {"years": YEARS, "through_week": through_week, "month": month,
            "generated": datetime.date.today().isoformat(), "metrics": metrics}
    through_short = f"Week {through_week}" if through_week else month
    data_through = f"Week {through_week} · {month} 2026" if through_week else f"{month} 2026"
    last_updated = f"Week {through_week} ({month} 2026)" if through_week else f"{month} 2026"
    insights = build_insights(metrics)
    ins_html = "".join(
        f'<li><span class="dot dot-{c}"></span><span>{t}</span></li>' for c, t in insights)
    html = (TEMPLATE
        .replace("__DATA__", json.dumps(data))
        .replace("__LAST_UPDATED__", last_updated)
        .replace("__DATA_THROUGH_PLAIN__", data_through.replace(" · ", ", "))
        .replace("__DATA_THROUGH__", data_through)
        .replace("__THROUGH_SHORT__", through_short)
        .replace("__INSIGHTS__", ins_html))
    os.makedirs(os.path.join(BASE, "data"), exist_ok=True)
    with open(os.path.join(BASE, "data", "vitals.json"), "w") as f:
        json.dump(data, f, indent=2)
    with open(os.path.join(BASE, "vitals.html"), "w") as f:
        f.write(html)
    return data_through

def apply_overlay(metrics, through_week, month):
    """Merge data/2026-current.json (the monthly Planning Center pull) over the
    current-year column. Leaves the frozen 2022-2025 history from the xlsx intact.
    Overlay schema:
      {"through_week": 30, "month": "July",
       "metrics": {"<metric name>": {"value": <num>, "pct": <num|null>,
                                     "total": <num, giving only>}, ...}}
    Any metric not listed keeps its spreadsheet value (carry-forward)."""
    path = os.path.join(BASE, "data", "2026-current.json")
    if not os.path.exists(path):
        return metrics, through_week, month, []
    ov = json.load(open(path))
    applied = []
    for m in metrics:
        o = ov.get("metrics", {}).get(m["name"])
        if not o:
            continue
        if "value" in o and o["value"] is not None:
            m["values"][-1] = o["value"]; applied.append(m["name"])
        if "pct" in o and m.get("pcts") is not None:
            m["pcts"][-1] = o["pct"]
        if "total" in o and m.get("totals") is not None:
            m["totals"][-1] = o["total"]
    return metrics, ov.get("through_week", through_week), ov.get("month", month), applied

def main():
    metrics, tw, month = extract()
    metrics, tw, month, applied = apply_overlay(metrics, tw, month)
    if applied:
        print(f"Applied monthly overlay to: {', '.join(applied)}")
    dt = render(metrics, tw, month)
    print(f"Built vitals.html + data/vitals.json — data through {dt}")
    print(f"  {len(metrics)} metrics, through week {tw}, close of {month}")

if __name__ == "__main__":
    main()
